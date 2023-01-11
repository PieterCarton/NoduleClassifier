from openpyxl import load_workbook

def write_result_to_file(file, result):
    file.write(
        f"""
        SLIDE: {result[0]}                          \n
        --- Nodules:                                \n
        --- --- Small: {result[1]}                  \n
        --- --- Medium: {result[2]}                 \n
        --- --- Large: {result[3]}                  \n
                                                    \n
        --- Total area: {result[4]} mm2             \n
                                                    \n
        """
    )


def write_results_to_disk(results, location):
    f = open(location, "w")
    
    f.write(
        f"""
        --- NODULE CLASSIFICATION RESULTS ---       \n
        Number of processed slides: {len(results)}  \n
                                                    \n
        Per slide breakdown:                        \n
        """
    )

    for result in results:
        write_result_to_file(f, result)

    f.close()

def find_row(slide, sheet):
    for row in range(5, 26):
        if sheet.cell(row, 2).value in slide:
            return row
    return 1

def calc_loss(result, actual):
    if result > actual:
        return result - actual
    else:
        return 2 * (actual - result)

def write_results_to_excel(results, iteration, location):
    wb = load_workbook(location)
    sheet = wb.worksheets[0]

    # find place to insert
    insertion_point = None
    col = 3
    while True:
        insertion_point = sheet.cell(2, col)
        if not insertion_point.value:
            break
        col += 4
    
    insertion_point.value = iteration

    # insert results into excel sheet
    for result in results:
        r = find_row(result[0], sheet)
        for i in range(3):
            sheet.cell(r, col + i).value = result[i + 1]
    
    large_loss = 0
    medium_loss = 0
    small_loss = 0
    num_results = 0

    # calculate loss function value
    for result in results:
        r = find_row(result[0], sheet)

        if not sheet.cell(r, 3).value:
            continue

        large_loss += calc_loss(result[1], sheet.cell(r, 3).value)
        medium_loss += calc_loss(result[2], sheet.cell(r, 4).value)
        small_loss += calc_loss(result[3], sheet.cell(r, 5).value)

        num_results += 1

    # insert loss into sheet
    sheet.cell(28, col).value = "Loss Function"
    sheet.cell(29, col).value = large_loss / num_results
    sheet.cell(29, col + 1).value = medium_loss / num_results
    sheet.cell(29, col + 2).value = small_loss / num_results

    wb.save(location)